import logging
import tempfile
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

from convert import convert, SHEET_ORDER

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

MAX_UPLOAD_MB = 15

st.set_page_config(page_title="Azure Cost Converter", page_icon="\u2601\ufe0f", layout="centered")

st.title("\u2601\ufe0f Azure Cost Estimation Converter")
st.write("Upload your Azure Calculator export to separate compute/license costs and fetch live Reserved Instance pricing.")

currency = st.selectbox("Select Target Currency:", ["INR", "USD", "EUR", "GBP", "AUD"])
st.caption("The original Calculator export's PAYG/RI figures are assumed to already be in this currency.")
uploaded_file = st.file_uploader("Upload Azure Export (.xlsx)", type=["xlsx"])


def _count_remarks(output_path):
    """Scan the generated workbook for any populated Remarks cells so the
    user is never left unaware that a row was approximated, degraded, or
    failed to process - instead of a bare 'Conversion Complete!' that hides
    row-level issues."""
    flagged = []
    wb = load_workbook(output_path, data_only=True)
    for sheet_name in SHEET_ORDER:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3, values_only=True):
            remark = row[8] if len(row) > 8 else None
            desc = row[4] if len(row) > 4 else ""
            if remark:
                flagged.append((sheet_name, desc, remark))
    return flagged


if uploaded_file is not None:
    size_mb = uploaded_file.size / (1024 * 1024)
    if size_mb > MAX_UPLOAD_MB:
        st.error(f"File is {size_mb:.1f} MB, which exceeds the {MAX_UPLOAD_MB} MB limit. Please split large estimates into smaller exports.")
    elif st.button("Convert File", type="primary"):
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / f"input_{uploaded_file.file_id}.xlsx"
            output_path = Path(temp_dir) / f"Processed_Estimate_{currency}.xlsx"

            with open(input_path, "wb") as f:
                f.write(uploaded_file.getbuffer())

            with st.spinner("Querying Azure Retail Pricing API & processing..."):
                try:
                    convert(str(input_path), str(output_path), currency)

                    with open(output_path, "rb") as f:
                        file_data = f.read()

                    flagged_rows = _count_remarks(output_path)
                    if flagged_rows:
                        st.warning(
                            f"Conversion complete, but {len(flagged_rows)} row(s) have notes in the "
                            f"**Remarks** column (approximated pricing, unmatched SKUs, or processing "
                            f"issues). Please review them in the downloaded file before using it as a final BOQ."
                        )
                        with st.expander("Show flagged rows"):
                            for sheet_name, desc, remark in flagged_rows:
                                st.markdown(f"- **{sheet_name}** — _{desc}_: {remark}")
                    else:
                        st.success("Conversion Complete! No rows required approximation or manual review.")

                    st.download_button(
                        label="\U0001f4e5 Download Processed Estimate",
                        data=file_data,
                        file_name=f"Processed_Estimate_{currency}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except ValueError as ve:
                    # ValueErrors raised by convert() are already user-facing,
                    # deliberately-worded validation messages - safe to show directly.
                    st.error(f"Validation Error: {ve}")
                except Exception as e:
                    # Never show raw internal exception text/paths to the user;
                    # log the full detail server-side for diagnosis instead.
                    log.error(f"Unhandled error converting '{uploaded_file.name}' (currency={currency}): {e}", exc_info=True)
                    st.error(
                        "An unexpected error occurred while processing this file. "
                        "Please confirm it's an unmodified Azure Pricing Calculator export and try again. "
                        "If the problem persists, contact support with the time of this attempt."
                    )
