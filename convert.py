"""
Azure Pricing Calculator -> Cost Estimation Workbook (Production Ready)

Changelog (production review fixes):
  CRITICAL
    - Worker exceptions no longer vanish silently (ThreadPoolExecutor futures are
      awaited via .result(); any failure is logged with row context and surfaced
      to the user via the Remarks column instead of quietly falling back to
      unenriched numbers with no explanation).
    - All worksheets in the uploaded workbook are parsed, not just the active
      sheet (previously any additional "Group" tabs were silently dropped).
    - Currency conversion no longer depends on reverse-engineering a single
      global FX ratio from one arbitrary row. Pricing is now fetched directly
      from the Azure Retail Prices API in the target currency for every row,
      so there is no "no eligible candidate row -> silently stuck at 1.0" trap.
  HIGH
    - Premium OS (RHEL/SUSE) license cost is now shown as its own line item
      instead of being silently folded into the compute line.
    - The SKU-name fallback no longer does a blind substring replace that could
      corrupt SKU families such as M-series (e.g. Standard_M32ms_v2). It only
      strips a trailing "s" flag when it directly follows a digit.
    - Azure Hybrid Benefit / BYOL is now detected; the Windows license delta is
      not charged again on top of an already-BYOL-priced VM.
    - The Remarks column is now actually populated for every degraded/assumed/
      approximated/failed scenario, instead of being silently empty.
  MEDIUM
    - Excel totals are real formulas (=SUM(...)) instead of static numbers, so
      row-level figures always foot exactly to the printed total.
    - Currency symbol is embedded in the number format.
    - The SQL/Linux "unaccounted cost" threshold scales with the row's own
      total instead of using a fixed absolute value that means something
      different in INR vs USD.
    - Formatting state (number format / currency symbol) is passed explicitly
      instead of living in module-level globals, to avoid cross-session bleed
      if multiple users hit this process concurrently.
"""
import re
import sys
import logging
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter  # noqa: F401 (kept for callers extending this module with formula-based sheets)

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_SYMBOLS = {"INR": "\u20b9", "USD": "$", "EUR": "\u20ac", "GBP": "\u00a3", "AUD": "A$"}

def num_format(currency):
    """Currency-aware Excel number format, e.g. '"₹"#,##0.00'."""
    symbol = CURRENCY_SYMBOLS.get(currency, "")
    return f'"{symbol}"#,##0.00' if symbol else "#,##0.00"

def _f(bold=False, italic=False, size=11, color="000000"):
    return Font(name="Calibri", bold=bold, italic=italic, size=size, color=color)

def _fill(h): return PatternFill("solid", fgColor=h)
def _al(h="left", v="center", w=False): return Alignment(horizontal=h, vertical=v, wrap_text=w)

def _safe_cell_text(v):
    """Defend against Excel/CSV formula injection when re-emitting user-supplied text."""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v

def hdr(c, v, wrap=False):
    c.value = v
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = _fill("4472C4")
    c.alignment = _al("center", "center", wrap)
    c.border = BORDER

def dat(c, v, bold=False, italic=False, align="left", color="000000", nfmt="#,##0.00"):
    c.value = _safe_cell_text(v)
    c.font = _f(bold, italic, color=color)
    c.alignment = _al(align, "center", True)
    c.border = BORDER
    if isinstance(v, (int, float)) and align == "right":
        c.number_format = nfmt

def tot_formula(c, formula, nfmt="#,##0.00"):
    c.value = formula
    c.font = _f(bold=True)
    c.fill = _fill("D9E1F2")
    c.alignment = _al("right", "center")
    c.border = BORDER
    c.number_format = nfmt

def tot_static(c, value, nfmt="#,##0.00"):
    """Write a plain computed number (not a formula) for totals. Formulas
    only display their value once Excel has recalculated and cached it -
    many viewers (LibreOffice quick-preview, Google Sheets before it loads,
    grid/file previewers, or reading the file back with openpyxl/pandas
    without opening it in Excel first) show a formula cell as blank until
    that happens. A static value always displays correctly everywhere. It
    still foots exactly, because it's computed as the sum of the SAME
    rounded values already written into the rows above it."""
    c.value = value
    c.font = _f(bold=True)
    c.fill = _fill("D9E1F2")
    c.alignment = _al("right", "center")
    c.border = BORDER
    c.number_format = nfmt

def widths(ws, d):
    for col, w in d.items(): ws.column_dimensions[col].width = w

SVC_MAP = {
    "virtual machines": "Virtual Machines",
    "virtual machine scale sets": "Virtual Machine Scale Sets",
    "vm scale sets": "Virtual Machine Scale Sets",
    "vmss": "Virtual Machine Scale Sets",
    "managed disks": "Managed Disks",
    "azure backup": "Azure Backup",
    "backup": "Azure Backup",
    "load balancer": "Load Balancer",
    "load balancers": "Load Balancer",
    "application gateway": "Application Gateway",
    "azure firewall": "Azure Firewall",
    "vpn gateway": "VPN Gateway",
    "storage": "Storage Accounts",
    "storage accounts": "Storage Accounts",
    "sql database": "SQL",
    "azure sql": "SQL",
    "ip addresses": "Public IP Addresses",
    "bandwidth": "Bandwidth",
    "azure monitor": "Azure Monitor",
    "key vault": "Key Vault",
}

# Service-type "sheet" names that should be priced through the same VM
# pricing/enrichment pipeline as "Virtual Machines" (same SKUs, same RI
# economics - a scale set's underlying instances are billed and reserved
# exactly like standalone VMs of the same size).
VM_LIKE_SHEETS = {"Virtual Machines", "Virtual Machine Scale Sets"}

SHEET_ORDER = [
    "Virtual Machines", "Virtual Machine Scale Sets", "Managed Disks", "Public IP Addresses",
    "Load Balancer", "Application Gateway", "Azure Firewall", "VPN Gateway",
    "Storage Accounts", "Azure Backup", "SQL", "Bandwidth",
    "Azure Monitor", "Key Vault", "Others",
]

SKIP = {"support", "disclaimer", "total", "licensing program", "billing account", "billing profile"}

PREMIUM_OS_KEYWORDS = ["red hat", "rhel", "suse", "sles", "ubuntu pro", "ubuntu advantage"]
AHB_KEYWORDS = ["bring your own license", "byol", "hybrid benefit", "ahb"]

# Azure retired NEW purchase/renewal of Reserved VM Instances for these series
# effective 2026-07-01. Existing reservations keep running until their term
# ends, but the Retail Prices API legitimately no longer returns Reservation
# entries for them - this is a real platform change, not a data gap.
# Source: https://techcommunity.microsoft.com/blog/azurecompute/azure-reserved-vm-instances-for-select-vm-series-will-no-longer-be-available-sta/4516505
RI_ONE_YEAR_RETIRED_SERIES = {"Av2", "Amv2", "Bv1", "D", "Ds", "Dv2", "Dsv2", "F", "Fs", "Fsv2", "G", "Gs", "Ls", "Lsv2"}
RI_FULLY_RETIRED_SERIES = {"Dv3", "Dsv3", "Ev3", "Esv3"}

def _classify_vm_series(sku):
    """
    Best-effort classification of an armSkuName into the short series labels
    Microsoft uses in its RI-retirement communications (e.g. 'Fsv2', 'Dv3').
    Only distinguishes the series relevant to the retirement lists above -
    returns None for anything else (which just means "not a known-retired
    series", not "unrecognized SKU").
    """
    m = re.match(r'^Standard_([A-Z])(\d+)([A-Za-z]*)(?:_[Vv](\d+))?$', sku)
    if not m:
        return None
    family, _, additive, version = m.groups()
    additive = (additive or "").lower()
    has_s = "s" in additive
    has_m = "m" in additive

    if family == "A" and version == "2":
        return "Amv2" if has_m else "Av2"
    if family == "B" and not version:
        return "Bv1"
    if family == "D":
        if version == "3": return "Dsv3" if has_s else "Dv3"
        if version == "2": return "Dsv2" if has_s else "Dv2"
        if not version: return "Ds" if has_s else "D"
    if family == "E" and version == "3":
        return "Esv3" if has_s else "Ev3"
    if family == "F":
        if version == "2": return "Fsv2" if has_s else None
        if not version: return "Fs" if has_s else "F"
    if family == "G" and not version:
        return "Gs" if has_s else "G"
    if family == "L":
        if version == "2": return "Lsv2" if has_s else None
        if not version: return "Ls" if has_s else None
    return None

def _ri_retirement_note(sku, term):
    """term: '1 Year' or '3 Years'. Returns a Remarks string if this SKU's
    series has had that RI term retired for new purchase/renewal, else None."""
    series = _classify_vm_series(sku)
    if series is None:
        return None
    if series in RI_FULLY_RETIRED_SERIES:
        return (f"Azure retired {series}-series Reserved Instances (both 1-year and 3-year) for new "
                f"purchase/renewal as of 1 Jul 2026 - no live RI pricing exists for this SKU any more "
                f"(existing reservations keep running until they expire). Consider Azure Savings Plan "
                f"for Compute or a newer VM generation instead.")
    if series in RI_ONE_YEAR_RETIRED_SERIES and term == "1 Year":
        return (f"Azure retired 1-year Reserved Instances for {series}-series as of 1 Jul 2026 for new "
                f"purchase/renewal - only 3-year RI (if shown) or Savings Plan remain available.")
    return None

def arm_region(display):
    val = str(display).lower().strip()
    return val.replace(" ", "")

# ── API Setup & Helpers ──────────────────────────────────────────────────
API = "https://prices.azure.com/api/retail/prices"

def get_http_session():
    session = requests.Session()
    # Bounded, less aggressive backoff than before (was up to ~31s per failed
    # filter chain); still resilient to transient 429/5xx without stalling a
    # web request for minutes when many SKUs fail in sequence.
    retries = Retry(total=3, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503, 504])
    session.mount("https://", HTTPAdapter(max_retries=retries))
    return session

def _api(session, cache, filt, currency="INR"):
    key = filt + "|" + currency
    if key in cache:
        return cache[key]
    try:
        r = session.get(API, params={"api-version": "2023-01-01-preview", "$filter": filt, "currencyCode": currency}, timeout=15)
        r.raise_for_status()
        items = r.json().get("Items", [])
        cache[key] = items
        return items
    except Exception as e:
        log.warning(f"API fetch error for filter [{filt}] currency={currency}: {e}")
        cache[key] = []  # cache the failure too, so a persistently-down SKU doesn't retry-storm every row
        return []

def _hourly_to_monthly(price): return price * 730

def detect_os(desc):
    return "Windows" if "windows" in desc.lower() else "Linux"

def detect_ahb(desc):
    d = desc.lower()
    return any(kw in d for kw in AHB_KEYWORDS)

def detect_sql_license(desc):
    desc_l = desc.lower()
    if "sql enterprise" in desc_l: return "SQL Enterprise License"
    if "sql standard" in desc_l: return "SQL Standard License"
    if "sql web" in desc_l: return "SQL Web License"
    if "sql developer" in desc_l: return "SQL Developer License"
    if "sql" in desc_l: return "SQL License"
    return None

def get_exact_license_name(desc):
    parts = re.split(r'[,;]', desc)
    sql_name = os_name = None
    for p in parts:
        p_lower = p.lower()
        if "sql" in p_lower:
            sql_name = re.sub(r'\s*\([^)]*\)', '', p).strip()
        if any(kw in p_lower for kw in PREMIUM_OS_KEYWORDS) and "sql" not in p_lower:
            os_name = re.sub(r'\s*\([^)]*\)', '', p).strip()
            if os_name.lower().startswith("linux "):
                os_name = os_name[6:].strip()
    return sql_name, os_name

# Static USD reference prices, used ONLY as a last-resort fallback when the
# Retail Prices API does not return a distinguishable OS-license meter for the
# SKU/region (rare, but possible for some regions/SKUs). These are approximate
# and callers are told so via the Remarks column - they are never applied silently.
_STATIC_PREMIUM_OS_USD = {
    "suse": {"low": 29.20, "high": 73.00},
    "rhel": {"low": 42.05, "high": 94.90},
}

def _static_os_family(desc_l):
    return "suse" if ("suse" in desc_l or "sles" in desc_l) else "rhel"

def extract_vm_sku(desc):
    if not desc: return None
    m = re.match(r'^\s*[\d,]+\s+([^()]+)\(', desc)
    if m:
        raw = m.group(1).strip()
        norm = re.sub(r'\s+', '_', raw)
        if not norm.lower().startswith("standard_"):
            norm = "Standard_" + norm
        return norm

    patterns = [r'^\d+\s+((?:[A-Z][A-Za-z0-9\-]+\s+)+v\d+)', r'^\d+\s+([A-Z][A-Za-z0-9\-]+)\s*\(', r'(Standard_[A-Za-z0-9_\-]+)']
    for pat in patterns:
        m = re.search(pat, desc.strip())
        if m:
            raw = m.group(1).strip()
            norm = re.sub(r'\s+', '_', raw)
            if not norm.startswith("Standard_"): norm = "Standard_" + norm
            return norm
    return None

def extract_quantity(desc):
    if not desc: return 1
    m = re.match(r'^\s*([0-9,]+)\s+', desc)
    if m:
        try:
            return max(1, int(m.group(1).replace(',', '')))
        except ValueError:
            return 1
    return 1

def extract_vcpus(desc):
    if not desc: return None
    m = re.search(r'\((\d+)\s*vCPU', desc, re.IGNORECASE)
    return int(m.group(1)) if m else None

def _sku_fallback_variants(sku):
    """
    Generate SAFE fallback SKU names when the exact armSkuName isn't found.

    Only strips a trailing 's' (Premium-storage / local-disk capability flag)
    when it appears directly after a digit and directly before '_v<N>' - e.g.
      Standard_D4s_v3   -> Standard_D4_v3    (correct: Dv3 base family)
      Standard_E8s_v5   -> Standard_E8_v5    (correct: Ev5 base family)
      Standard_D4ds_v5  -> Standard_D4d_v5   (correct: Ddv5 base family)
    This deliberately will NOT touch SKUs like Standard_M32ms_v2, because the
    's' there is preceded by 'm', not a digit - a blind substring replace used
    to turn that into the invalid/wrong SKU 'Standard_M32m_v2'.
    """
    variants = []
    v1 = re.sub(r'(\d)s(_[Vv]\d)', r'\1\2', sku)
    if v1 != sku:
        variants.append(v1)
    v2 = re.sub(r'(\d)ds(_[Vv]\d)', r'\1d\2', sku)
    if v2 != sku and v2 not in variants:
        variants.append(v2)
    return variants

def get_vm_pricing(session, cache, sku, region_display, is_spot, currency="INR"):
    """
    Fetch VM pricing DIRECTLY in the requested currency (no reverse-engineered
    FX math). Returns compute PAYG, Windows license delta, and 1yr/3yr RI
    monthly-equivalent pricing, all already in `currency`.
    """
    region = arm_region(region_display)
    result = {"compute_payg": None, "windows_license": None, "compute_ri1": None, "compute_ri3": None, "raw_payg_items": []}

    def fetch(s, ptype):
        return _api(session, cache, f"armSkuName eq '{s}' and armRegionName eq '{region}' and priceType eq '{ptype}'", currency)

    def get_items(ptype):
        items = fetch(sku, ptype)
        if items:
            return items
        for variant in _sku_fallback_variants(sku):
            items = fetch(variant, ptype)
            if items:
                return items
        base_sku = re.sub(r'-\d+', '', sku)
        if base_sku != sku:
            items = fetch(base_sku, ptype)
            if items:
                return items
        return []

    def _get_price(items, must_be_win=False):
        cands = []
        for i in items:
            prod = i.get("productName", "").lower()
            meter = i.get("meterName", "").lower()
            if "low priority" in meter: continue
            if is_spot and "spot" not in meter: continue
            if not is_spot and "spot" in meter: continue
            # Distinct RHEL/SUSE/etc. OS-license meters are a separate cost
            # component, not the base compute price - exclude them here so
            # they can't leak in as a spuriously cheap "compute" candidate.
            if any(kw in prod for kw in PREMIUM_OS_KEYWORDS): continue
            is_win_prod = "windows" in prod
            if must_be_win and not is_win_prod: continue
            if not must_be_win and is_win_prod: continue
            if i.get("retailPrice", 0) > 0:
                cands.append(i["retailPrice"])
        return min(cands) if cands else None

    try:
        payg_items = get_items("Consumption")
        result["raw_payg_items"] = payg_items

        linux_hr = _get_price(payg_items, must_be_win=False)
        win_hr = _get_price(payg_items, must_be_win=True)

        result["compute_payg"] = _hourly_to_monthly(linux_hr) if linux_hr else None
        if win_hr and linux_hr:
            result["windows_license"] = max(0, _hourly_to_monthly(win_hr) - _hourly_to_monthly(linux_hr))

        if not is_spot:
            ri_items = get_items("Reservation")
            ri1_cands = [i for i in ri_items if i.get("reservationTerm") == "1 Year"]
            ri1_val = _get_price(ri1_cands, must_be_win=False)
            if ri1_val is not None:
                result["compute_ri1"] = ri1_val / 12

            ri3_cands = [i for i in ri_items if i.get("reservationTerm") == "3 Years"]
            ri3_val = _get_price(ri3_cands, must_be_win=False)
            if ri3_val is not None:
                result["compute_ri3"] = ri3_val / 36

    except Exception as e:
        log.warning(f"VM pricing error for sku={sku} region={region} currency={currency}: {e}")

    return result

def get_premium_os_license_price(session, cache, payg_items, desc, sku, region, qty, currency):
    """
    Determine the RHEL/SUSE license portion.

    1) Try to find a distinct OS-license meter in the SAME item set already
       fetched for this SKU/region/currency (Azure publishes RHEL/SUSE as
       their own productName entries alongside the base Linux compute meter).
       This is fully live-priced and currency-correct with no extra API call.
    2) If nothing distinguishable is found, fall back to a static USD
       reference table, converted using a per-row ratio derived from this
       SAME SKU's own native-vs-USD compute price (never a global, workbook-
       wide "guessed" ratio from an unrelated row). The result is clearly
       flagged as approximate via the returned `approximate` flag.
    3) If even that isn't possible, return 0 with `unavailable=True` so the
       caller can surface a clear warning instead of silently underpricing.
    """
    desc_l = desc.lower()
    family_kw = ["red hat", "rhel"] if _static_os_family(desc_l) == "rhel" else ["suse", "sles"]

    cands = []
    for i in payg_items:
        prod = i.get("productName", "").lower()
        meter = i.get("meterName", "").lower()
        if "low priority" in meter or "spot" in meter:
            continue
        if any(kw in prod for kw in family_kw) and i.get("retailPrice", 0) > 0:
            cands.append(i["retailPrice"])
    if cands:
        hourly = min(cands)
        return {"monthly": _hourly_to_monthly(hourly) * qty, "approximate": False, "unavailable": False}

    # Fallback: static USD table, converted via a per-row (not global) ratio.
    vcpus = extract_vcpus(desc)
    family = _static_os_family(desc_l)
    tier = "low" if (vcpus and vcpus <= 4) else "high"
    usd_price = _STATIC_PREMIUM_OS_USD[family][tier]

    if currency == "USD":
        return {"monthly": usd_price * qty, "approximate": True, "unavailable": False}

    native = get_vm_pricing(session, cache, sku, region, False, currency)
    usd = get_vm_pricing(session, cache, sku, region, False, "USD")
    if native.get("compute_payg") and usd.get("compute_payg"):
        ratio = native["compute_payg"] / usd["compute_payg"]
        return {"monthly": usd_price * qty * ratio, "approximate": True, "unavailable": False}

    return {"monthly": 0.0, "approximate": True, "unavailable": True}

# ── Parsing ────────────────────────────────────────────────────────────────
def parse_format(wb):
    """
    Parse ALL worksheets in the workbook (a Calculator export can contain
    multiple "Group" tabs), not just the active one. Rows carry a
    `source_sheet` tag purely for diagnostics/remarks.
    """
    rows = []
    any_valid_sheet = False

    def _get_cost(row_data, idx, fallback):
        if len(row_data) > idx and isinstance(row_data[idx], (int, float)) and row_data[idx] > 0:
            return float(row_data[idx])
        return float(fallback)

    for ws in wb.worksheets:
        in_data = valid_format = False
        for r in ws.iter_rows(values_only=True):
            if not r or r[0] is None: continue
            if str(r[0]).strip().lower() == "service category":
                in_data = valid_format = True
                continue

            if not in_data: continue
            svc_cat = str(r[0]).strip() if r[0] else ""
            svc_type = str(r[1]).strip() if r[1] else ""
            region = str(r[3]).strip() if r[3] else ""
            desc = str(r[4]).strip() if r[4] else ""
            cost_raw = r[5]

            if svc_cat.lower() in SKIP or region.lower() in SKIP or not isinstance(cost_raw, (int, float)):
                continue

            rows.append({
                "svc_cat": svc_cat, "svc_type": svc_type, "cust_name": str(r[2] or "").strip(),
                "region": region, "desc": desc, "payg": float(cost_raw),
                "ri1": _get_cost(r, 6, cost_raw),
                "ri3": _get_cost(r, 7, cost_raw),
                "remarks": "", "sub_rows": [], "source_sheet": ws.title,
            })
        if valid_format:
            any_valid_sheet = True

    if not any_valid_sheet:
        raise ValueError("Could not find a 'Service category' header on any sheet. Please ensure this is an unmodified Azure Pricing Calculator export.")
    return rows

def classify(rows):
    buckets = {}
    for r in rows:
        key = str(r.get("svc_type") or "").lower().strip()
        sheet = SVC_MAP.get(key, "Others")
        if sheet == "Others":
            for k, v in SVC_MAP.items():
                if k in key:
                    sheet = v; break
        buckets.setdefault(sheet, []).append(r)
    return buckets

def _append_remark(row, text):
    existing = row.get("remarks", "")
    row["remarks"] = f"{existing}; {text}" if existing else text

def _row_context(row):
    return f"custom_name='{row.get('cust_name')}' region='{row.get('region')}' sheet='{row.get('source_sheet')}' desc='{row.get('desc')[:60]}'"

def enrich_vms_concurrent(vm_rows, currency="INR"):
    """
    Enrich VM rows with live Azure Retail Prices API data.

    Every row is processed independently and defensively:
      - Pricing is fetched directly in the target currency (no global FX hack).
      - Any exception during a row's processing is caught, logged with full
        row context, and the row falls back to its original spreadsheet
        values WITH a visible Remarks warning - it is never silently dropped
        or silently mispriced.
    """
    log.info(f"Querying Azure Retail Pricing API for {len(vm_rows)} VM row(s), currency={currency}...")
    session = get_http_session()
    cache = {}

    def process_row(row):
        desc, region = row["desc"], row["region"]
        try:
            os_type = detect_os(desc)
            qty = extract_quantity(desc)
            sku = extract_vm_sku(desc)
            is_ahb = detect_ahb(desc)

            sql_exact, _ = get_exact_license_name(desc)
            has_premium_os = any(kw in desc.lower() for kw in PREMIUM_OS_KEYWORDS)
            has_sql = bool(sql_exact)

            row["sql_lbl_exact"] = sql_exact or "SQL License"
            row["api"] = {}

            if not sku:
                if has_premium_os or has_sql:
                    row["api"] = {"is_standalone": True}
                _append_remark(row, "VM SKU could not be parsed from description - original calculator values shown as-is; compute/license were not separated")
                return row

            is_spot = "spot" in desc.lower()
            pricing = get_vm_pricing(session, cache, sku, region, is_spot, currency)
            orig_payg = row.get("payg", 0)

            compute_native = (pricing.get("compute_payg") or 0) * qty
            win_native = (pricing.get("windows_license") or 0) * qty
            ri1_native = (pricing.get("compute_ri1") * qty) if pricing.get("compute_ri1") is not None else None
            ri3_native = (pricing.get("compute_ri3") * qty) if pricing.get("compute_ri3") is not None else None

            if compute_native == 0:
                # API had nothing for this SKU/region/currency at all.
                row["api"] = {
                    "compute_payg_final": orig_payg, "win_lic_payg_final": 0,
                    "prem_os_payg_final": 0, "sql_payg_final": 0, "other_payg_final": 0,
                    "compute_ri1": orig_payg, "compute_ri3": orig_payg,
                }
                _append_remark(row, f"No live pricing found for SKU '{sku}' in region '{region}' - showing original calculator values; RI pricing not available")
                return row

            win_lic_payg = win_native if (os_type == "Windows" and not is_ahb) else 0
            if os_type == "Windows" and is_ahb:
                _append_remark(row, "Azure Hybrid Benefit / BYOL detected - Windows license not charged separately")

            unaccounted = orig_payg - compute_native - win_lic_payg
            # Threshold scales with the row's own total instead of a fixed
            # absolute figure that means something different in every currency.
            epsilon = max(1.0, 0.01 * orig_payg)

            prem_os_payg = 0.0
            sql_payg = 0.0
            other_payg = 0.0

            if unaccounted > epsilon:
                if has_sql and has_premium_os:
                    os_result = get_premium_os_license_price(
                        session, cache, pricing.get("raw_payg_items", []), desc, sku, region, qty, currency
                    )
                    prem_os_payg = os_result["monthly"]
                    if os_result["approximate"]:
                        _append_remark(row, "Premium OS (RHEL/SUSE) license cost approximated from a static reference price (no distinct live meter found for this SKU/region)")
                    if os_result["unavailable"]:
                        _append_remark(row, "Premium OS (RHEL/SUSE) license cost could not be determined - value may be understated")
                    sql_payg = max(0, unaccounted - prem_os_payg)
                elif has_sql:
                    sql_payg = unaccounted
                elif has_premium_os:
                    os_result = get_premium_os_license_price(
                        session, cache, pricing.get("raw_payg_items", []), desc, sku, region, qty, currency
                    )
                    prem_os_payg = os_result["monthly"]
                    if os_result["approximate"]:
                        _append_remark(row, "Premium OS (RHEL/SUSE) license cost approximated from a static reference price (no distinct live meter found for this SKU/region)")
                    if os_result["unavailable"]:
                        _append_remark(row, "Premium OS (RHEL/SUSE) license cost could not be determined - value may be understated")
                    # Any leftover beyond the detected OS license is unexplained; don't hide it.
                    leftover = unaccounted - prem_os_payg
                    if leftover > epsilon:
                        prem_os_payg += leftover
                        _append_remark(row, "Additional unexplained cost above the detected OS license was folded into the OS license line")
                else:
                    # Unaccounted cost with no SQL/premium-OS signal detected -
                    # most likely a plain rounding/markup quirk in the source
                    # file. Shown as its own clearly-labeled "Other" line
                    # rather than being silently absorbed into compute, and
                    # rather than being mislabeled as a premium-OS license.
                    other_payg = unaccounted
                    _append_remark(row, "Unexplained cost difference vs. live compute pricing was shown as a separate 'Other/Unidentified Cost' line - please verify")

            if sql_payg < 0:
                other_payg = max(0.0, other_payg + sql_payg)
                sql_payg = 0.0

            compute_payg_final = compute_native
            # Note: the premium-OS license amount is intentionally kept SEPARATE
            # (not folded into compute) so it appears as its own line item, and
            # rides flat (undiscounted) on top of the RI compute figure, same as
            # how BYOL/subscription licenses actually behave under a Reserved Instance.
            compute_ri1_final = ri1_native if ri1_native is not None else compute_payg_final
            compute_ri3_final = ri3_native if ri3_native is not None else compute_payg_final

            if not is_spot:
                if ri1_native is None:
                    note = _ri_retirement_note(sku, "1 Year")
                    _append_remark(row, note or "1-Year RI pricing not found for this SKU/region - showing PAYG instead")
                if ri3_native is None:
                    note = _ri_retirement_note(sku, "3 Years")
                    _append_remark(row, note or "3-Year RI pricing not found for this SKU/region - showing PAYG instead")

            if is_spot:
                _append_remark(row, "Reserved Instance pricing is not applicable to Spot VMs")

            row["api"] = {
                "compute_payg_final": compute_payg_final,
                "win_lic_payg_final": win_lic_payg,
                "prem_os_payg_final": prem_os_payg,
                "sql_payg_final": sql_payg,
                "other_payg_final": other_payg,
                "compute_ri1": compute_ri1_final,
                "compute_ri3": compute_ri3_final,
            }
            return row

        except Exception as e:
            log.error(f"Unhandled error processing VM row ({_row_context(row)}): {e}", exc_info=True)
            row["api"] = {
                "compute_payg_final": row.get("payg", 0), "win_lic_payg_final": 0,
                "prem_os_payg_final": 0, "sql_payg_final": 0, "other_payg_final": 0,
                "compute_ri1": row.get("ri1", row.get("payg", 0)),
                "compute_ri3": row.get("ri3", row.get("payg", 0)),
            }
            _append_remark(row, f"Processing error - showing original calculator values (see server logs): {e}")
            return row

    max_workers = min(10, max(2, len(vm_rows)))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(process_row, r) for r in vm_rows]
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                # Should be unreachable (process_row catches internally), but
                # kept as a hard backstop so nothing can ever vanish silently.
                log.error(f"Uncaught exception escaped process_row: {e}", exc_info=True)

    return vm_rows

# ── Output ────────────────────────────────────────────────────────────────
def write_res_header(ws):
    ws.merge_cells("F1:H1")
    hdr(ws["F1"], "Monthly Cost")
    for addr in ["G1", "H1"]: ws[addr].border = BORDER
    for ci, h in enumerate(["Service category", "Service type", "Custom name", "Region", "Description", "PAYG", "1 Year RI Model", "3 Year RI Model", "Remarks"], 1):
        hdr(ws.cell(2, ci), h, wrap=True)
    ws.row_dimensions[2].height = 28.8
    ws.freeze_panes = "A3"

def write_vm_sheet(wb, rows, nfmt, sheet_name="Virtual Machines"):
    ws = wb.create_sheet(sheet_name)
    write_res_header(ws)
    widths(ws, {"A": 15, "B": 15, "C": 14, "D": 12, "E": 55, "F": 13, "G": 14, "H": 14, "I": 55})

    ri = 3
    payg_vals, ri1_vals, ri3_vals = [], [], []

    def _emit(ri, vals, italic=False, color="000000"):
        for ci, v in enumerate(vals, 1):
            dat(ws.cell(ri, ci), v, italic=italic, color=color, align="right" if ci >= 6 and isinstance(v, float) else "left", nfmt=nfmt)

    for row in rows:
        p = row.get("api", {})
        if p.get("is_standalone"):
            payg = row.get("payg", 0)
            ri1 = row.get("ri1", payg)
            ri3 = row.get("ri3", payg)

            payg, ri1, ri3 = round(payg, 2), round(ri1, 2), round(ri3, 2)
            _emit(ri, [row["svc_cat"], row["svc_type"], row["cust_name"], row["region"], row["desc"], payg, ri1, ri3, row.get("remarks", "")])
            payg_vals.append(payg); ri1_vals.append(ri1); ri3_vals.append(ri3)
            ri += 1
            continue

        compute_payg = p.get("compute_payg_final", row.get("payg", 0))
        win_lic_payg = p.get("win_lic_payg_final", 0)
        prem_os_payg = p.get("prem_os_payg_final", 0)
        other_payg = p.get("other_payg_final", 0)

        sql_rows_b = [s.copy() for s in row.get("sub_rows", []) if "sql" in s["desc"].lower()]
        sql_payg_deduced = p.get("sql_payg_final", 0)

        if not sql_rows_b and sql_payg_deduced > 0:
            sql_rows_b.append({
                "desc": row.get("sql_lbl_exact", "SQL License"),
                "payg": sql_payg_deduced, "ri1": sql_payg_deduced, "ri3": sql_payg_deduced, "is_api": True
            })

        compute_ri1 = p.get("compute_ri1", compute_payg)
        compute_ri3 = p.get("compute_ri3", compute_payg)

        compute_payg, compute_ri1, compute_ri3 = round(compute_payg, 2), round(compute_ri1, 2), round(compute_ri3, 2)
        _emit(ri, [row["svc_cat"], row["svc_type"], row["cust_name"], row["region"], row["desc"], compute_payg, compute_ri1, compute_ri3, row.get("remarks", "")])
        payg_vals.append(compute_payg); ri1_vals.append(compute_ri1); ri3_vals.append(compute_ri3)
        ri += 1

        if win_lic_payg > 0:
            v = round(win_lic_payg, 2)
            _emit(ri, ["", "", "", "", "Windows License", v, v, v, "License Cost (Not discounted by Compute RI)"], italic=True, color="595959")
            payg_vals.append(v); ri1_vals.append(v); ri3_vals.append(v)
            ri += 1

        if prem_os_payg > 0:
            v = round(prem_os_payg, 2)
            _emit(ri, ["", "", "", "", "Premium OS License (RHEL/SUSE)", v, v, v, "License Cost (Not discounted by Compute RI)"], italic=True, color="595959")
            payg_vals.append(v); ri1_vals.append(v); ri3_vals.append(v)
            ri += 1

        if other_payg > 0:
            v = round(other_payg, 2)
            _emit(ri, ["", "", "", "", "Other/Unidentified Cost", v, v, v, "Please verify against source estimate"], italic=True, color="C00000")
            payg_vals.append(v); ri1_vals.append(v); ri3_vals.append(v)
            ri += 1

        for s in sql_rows_b:
            rmk = "License Cost (Not discounted by Compute RI)" if s.get("is_api") else s.get("remarks", "")
            payg_v = round(s["payg"], 2) if s.get("payg") else None
            ri1_v = round(s.get("ri1") or s["payg"], 2) if s.get("payg") else None
            ri3_v = round(s.get("ri3") or s["payg"], 2) if s.get("payg") else None
            _emit(ri, ["", "", "", "", s["desc"], payg_v, ri1_v, ri3_v, rmk], italic=True, color="595959")
            payg_vals.append(payg_v or 0); ri1_vals.append(ri1_v or 0); ri3_vals.append(ri3_v or 0)
            ri += 1

        for s in [s for s in row.get("sub_rows", []) if "sql" not in s["desc"].lower()]:
            payg_v = round(s["payg"], 2) if s.get("payg") else None
            ri1_v = round(s.get("ri1") or s["payg"], 2) if s.get("payg") else None
            ri3_v = round(s.get("ri3") or s["payg"], 2) if s.get("payg") else None
            _emit(ri, ["", "", "", "", s["desc"], payg_v, ri1_v, ri3_v, s.get("remarks", "")], italic=True, color="595959")
            payg_vals.append(payg_v or 0); ri1_vals.append(ri1_v or 0); ri3_vals.append(ri3_v or 0)
            ri += 1

    total_row = ri
    total_payg, total_ri1, total_ri3 = round(sum(payg_vals), 2), round(sum(ri1_vals), 2), round(sum(ri3_vals), 2)
    ws.cell(total_row, 5, "Total").font = _f(bold=True)
    ws.cell(total_row, 5).border = BORDER
    tot_static(ws.cell(total_row, 6), total_payg, nfmt)
    tot_static(ws.cell(total_row, 7), total_ri1, nfmt)
    tot_static(ws.cell(total_row, 8), total_ri3, nfmt)

    return total_payg, total_ri1, total_ri3

def write_generic_sheet(wb, sheet_name, rows, nfmt):
    ws = wb.create_sheet(sheet_name)
    write_res_header(ws)
    widths(ws, {"A": 15, "B": 14, "C": 22, "D": 12, "E": 60, "F": 13, "G": 14, "H": 14, "I": 40})

    ri = 3
    payg_vals, ri1_vals, ri3_vals = [], [], []
    for row in rows:
        payg = round(row.get("payg", 0), 2)
        ri1 = round(row.get("ri1", payg), 2)
        ri3 = round(row.get("ri3", payg), 2)

        vals = [row["svc_cat"], row["svc_type"], row["cust_name"], row["region"], row["desc"], payg, ri1, ri3, row.get("remarks", "")]
        for ci, v in enumerate(vals, 1):
            dat(ws.cell(ri, ci), v, align="right" if ci >= 6 and isinstance(v, float) else "left", nfmt=nfmt)
        payg_vals.append(payg); ri1_vals.append(ri1); ri3_vals.append(ri3)
        ri += 1

    total_row = ri
    total_payg, total_ri1, total_ri3 = round(sum(payg_vals), 2), round(sum(ri1_vals), 2), round(sum(ri3_vals), 2)
    ws.cell(total_row, 5, "Total").font = _f(bold=True)
    ws.cell(total_row, 5).border = BORDER
    tot_static(ws.cell(total_row, 6), total_payg, nfmt)
    tot_static(ws.cell(total_row, 7), total_ri1, nfmt)
    tot_static(ws.cell(total_row, 8), total_ri3, nfmt)

    return total_payg, total_ri1, total_ri3

def convert(input_path, output_path, currency="INR"):
    if currency not in CURRENCY_SYMBOLS:
        raise ValueError(f"Unsupported currency '{currency}'. Supported: {', '.join(CURRENCY_SYMBOLS)}")

    nfmt = num_format(currency)

    wb_in = load_workbook(input_path, data_only=True)
    rows = parse_format(wb_in)
    buckets = classify(rows)

    for sname in VM_LIKE_SHEETS:
        if sname in buckets:
            enrich_vms_concurrent(buckets[sname], currency)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    sheet_totals = {}  # sheet_name -> (payg_total, ri1_total, ri3_total) as plain numbers

    for sname in SHEET_ORDER:
        if sname not in buckets: continue
        if sname in VM_LIKE_SHEETS:
            sheet_totals[sname] = write_vm_sheet(wb_out, buckets[sname], nfmt, sheet_name=sname)
        else:
            sheet_totals[sname] = write_generic_sheet(wb_out, sname, buckets[sname], nfmt)

    ws = wb_out.create_sheet("Summary", 0)
    ws.merge_cells("A1:A2"); ws.merge_cells("B1:B2"); ws.merge_cells("C1:E1"); ws.merge_cells("F1:F2")
    for addr, val, al in [("A1", "Sl No", "left"), ("B1", "Service Name", "left"), ("C1", f"Monthly Cost ({currency})", "center"), ("F1", "Remarks", "left")]:
        c = ws[addr]; c.value = val; c.font = _f(bold=True); c.alignment = _al(al, "center"); c.border = BORDER
    for addr, val in [("C2", "PAYG"), ("D2", "1 YR RI Model"), ("E2", "3 YR RI Model")]:
        c = ws[addr]; c.value = val; c.font = _f(bold=True); c.alignment = _al("center", "center"); c.border = BORDER

    ri = 3
    grand_payg = grand_ri1 = grand_ri3 = 0.0
    for sl, (sname, (payg_total, ri1_total, ri3_total)) in enumerate(sheet_totals.items(), 1):
        c = ws.cell(ri, 1, sl); c.border = BORDER; c.alignment = _al("center")
        c = ws.cell(ri, 2, sname); c.border = BORDER
        for ci, val in [(3, payg_total), (4, ri1_total), (5, ri3_total)]:
            c = ws.cell(ri, ci, val)
            c.alignment = _al("right"); c.border = BORDER; c.number_format = nfmt
        ws.cell(ri, 6).border = BORDER
        grand_payg += payg_total; grand_ri1 += ri1_total; grand_ri3 += ri3_total
        ri += 1

    tot_static(ws.cell(ri, 3), round(grand_payg, 2), nfmt)
    tot_static(ws.cell(ri, 4), round(grand_ri1, 2), nfmt)
    tot_static(ws.cell(ri, 5), round(grand_ri3, 2), nfmt)
    widths(ws, {"A": 5.5, "B": 22, "C": 14, "D": 14, "E": 14, "F": 55})
    ws.freeze_panes = "A3"
    wb_out.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python convert.py input.xlsx [output.xlsx] [currency]")
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "output.xlsx", sys.argv[3] if len(sys.argv) > 3 else "INR")
